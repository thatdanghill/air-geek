import sys, getopt
from openpyxl import Workbook, load_workbook
import os
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'air_geek.settings')

import django
django.setup()

from django.core.exceptions import ObjectDoesNotExist
from django.contrib.auth.models import User
from miner.models import UserProfile, Project, Page, Graph, Point

def parseTable():
    try:
        user = User.objects.get(username = 'super')
        userprofile = UserProfile.objects.get(user = user)
        file_name = sys.argv[1]
        wb = load_workbook(file_name)
        project_name = sys.argv[2]
        project = Project.objects.get(user=userprofile, name=project_name)
        page_name = sys.argv[3]
        page = Page.objects.get(project = project, name = page_name)
        graph_name = sys.argv[4]
        graph = Graph.objects.get(page = page, name = graph_name)
        points = graph.points.all()
        worksheet = wb.active
        
        parseSheet(worksheet, points, graph)
    
    except User.DoesNotExist:
        print("User is not valid")
    except UserProfile.DoesNotExist:
        print("UserProfile is not valid")
    except getopt.GetoptError:
        print("Script requires 3 arguments: project name, page name, and graph name")
        sys.exit(2)
    except Project.DoesNotExist:
        print("Project does not exist!")
    except Page.DoesNotExist:
        print("Page does not exist!")
    except Graph.DoesNotExist:
        print("Graph does not exist!")

def parseSheet(ws, points, graph):
    i = 2
    j = 2
    while ws.cell(row = 1, column = j).value != None:
        while ws.cell(row = i, column = 1).value != None:
            if ws.cell(row = i, column = j).value != None:
                placeInOrder(ws.cell(row = i, column = 1).value, ws.cell(row = 1, column = j).value, ws.cell(row = i, column = j).value, points, graph)
            i += 1
        i = 2
        j += 1

def placeInOrder(month, year, y, points, graph):
    
    x = month + " " + str(year)
    
    if points.count() == 0:
        p = Point.objects.create(index = 0, x = x, y = float(y), graph= graph)
        p.save()
        return
    
    ind = findIndex(year, month, points)
    
    for i in range(ind, points.count()):
        q = points[i]
        q.index += 1
        q.save()
    
    p = Point.objects.create(index = ind, x = x, y = float(y), graph= graph)
    p.save()

def findIndex(year, month, points):
    moDic = ['jan', 'feb', 'mar', 'apr', 'may', 'jun', 'jul', 'aug', 'sep', 'oct', 'nov', 'dec', 'january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december']
    i = 0
    if points.count() == 0:
        print("This shouldn't happen")
        return 0
    
    while i < points.count() and int(points[i].x.split(" ")[1]) < year:
        i = i + 1
    
    while i < points.count() and int(points[i].x.split(" ")[1]) == year and moDic.index(points[i].x.split(" ")[0].lower()) % 12 < moDic.index(month.lower()) % 12:
        i = i + 1
    return i


if __name__ == '__main__':
    parseTable()